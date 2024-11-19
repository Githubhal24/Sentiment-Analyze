import os
import time
import openai
#import 
from openpyxl import Workbook
from dotenv import load_dotenv

#.envファイルからロード
load_dotenv()

openai.api_type = os.getenv("API_TYPE")
openai.api_base = os.getenv("API_BASE")
openai.api_version = os.getenv("API_VERSION")
openai.api_key = os.getenv("API_KEY")

#config
datafolder = r"C:User\username\work\"
filename = "data.xlsx"
sheetname = "Sheet"
column = 4
startRow = 2
endRow = 230
outputfolder = r"C:User\username\work\"
outputfilename = f"Analyzed_{filename}"

wb = openpyxl.load_workbook(f"{datafolder}{filename}")
ws = wb[sheetname]

#出力ファイルにコピー
outputworkbook = Workbook()
outputWs = outputworkbook.active
for row in ws.iter_rows():
    for cell in row:
        outputWs.cell(cell.row, cell.column, cell.value)

#列名を設定
outputWs.cell(row=1, column=1, value="No")
outputWs.cell(row=1, column=2, value="地方公共コード")
outputWs.cell(row=1, column=3, value="地方公共団体名")
outputWs.cell(row=1, column=4, value="内容")
outputWs.cell(row=1, column=5, value="文章")
outputWs.cell(row=1, column=6, value="文章パラメータ")
outputWs.cell(row=1, column=7, value="単語")
outputWs.cell(row=1, column=8, value="単語パラメータ")
outputWs.cell(row=1, column=9, value="注目単語１")
outputWs.cell(row=1, column=10, value="注目単語２")
outputWs.cell(row=1, column=11, value="総合")
outputWs.cell(row=1, column=12, value="総合パラメータ")

#感情分析
for i in range(startRow, endRow + 1):
    text = ws.cell(row=i, column=column).value
    if text is None or text.strip() == "":
        continue
    else:
        text = text.strip()   #空白の削除
        text = text.replace("\n", "") #改行の削除  

    content = f"""
    【要件】
    ・以下のアンケート結果を分析して、「ポジティブ」「ネガティブ」「ニュートラル」の３択の感情で回答してください。
    「ポジティブ」「ネガティブ」「ニュートラル」以外の感情を出力しないでください。

    ・分析は以下に示す２つの分析手法で実践してください。
    （１）文章全体から判断して「ポジティブ」「ネガティブ」「ニュートラル」の３択の感情で回答する。
    （２）文章を単語ごとに区切って、単語ごとの「ポジティブ」「ネガティブ」「ニュートラル」の３択の感情を考慮して回答する。
    ・選択した「ポジティブ」「ネガティブ」「ニュートラル」についてのパラメータを回答してください。
    パラメータは０～１０の範囲を取り、値が大きいほど選択した感情の傾向が高くなります。
    以下に回答すべきこと、回答例を示すので回答形式を順守してください。
    判断根拠は不要なので、回答例で示す回答フォーマットに従って「回答：」以下に回答結果のみ列挙していってください。

    【回答すべきこと】
    （手法（１）で「ポジティブ」「ネガティブ」「ニュートラル」のいずれかを選択した感情）
    （手法（１）で「ポジティブ」「ネガティブ」「ニュートラル」のいずれかを選択した感情のパラメータ）
    （手法（２）で「ポジティブ」「ネガティブ」「ニュートラル」のいずれかを選択した感情）
    （手法（２）で「ポジティブ」「ネガティブ」「ニュートラル」のいずれかを選択した感情のパラメータ）
    （手法（２）で「ポジティブ」「ネガティブ」「ニュートラル」のいずれかを選択した感情を選択するにあたって１番目に注目した単語）
    （手法（２）で「ポジティブ」「ネガティブ」「ニュートラル」のいずれかを選択した感情を選択するにあたって２番目に注目した単語）
    （手法（１）と手法（２）から総合的に判断した「ポジティブ」「ネガティブ」「ニュートラル」のいずれかを選択した感情）
    （手法（１）と手法（２）から総合的に判断した「ポジティブ」「ネガティブ」「ニュートラル」のいずれかを選択した感情のパラメータ）


    【回答例】
    ポジティブ
    10  
    ニュートラル
    5
    普通
    平穏
    ポジティブ
    8

    文章：
    {text}
    回答：
    """

    try:
        print(f"{i}: start")
        response = openai.ChatCompletion.create(
            deployment_id = "gpt-4o",
            max_token = 40,
            stop = ["」"],
            message = [
                {"role": "system", "content": "あなたは感情分析のプロです。"},
                {"role": "system", "content": content},                
            ], 
        )

        result = response["choice"][0]["message"]["content"]
        print(f"{i}: {result}")


        lines = result.split("\n")
        outputWs.cell(row=i, column=1, value=ws.cell(row=i, column=1).value)
        outputWs.cell(row=i, column=2, value=ws.cell(row=i, column=2).value)
        outputWs.cell(row=i, column=3, value=ws.cell(row=i, column=3).value)
        outputWs.cell(row=i, column=4, value=text)
        outputWs.cell(row=i, column=5, value=lines[0].strip())
        outputWs.cell(row=i, column=6, value=lines[1].strip())
        outputWs.cell(row=i, column=7, value=lines[2].strip())
        outputWs.cell(row=i, column=8, value=lines[3].strip())
        outputWs.cell(row=i, column=9, value=lines[4].strip())
        outputWs.cell(row=i, column=10, value=lines[5].strip())
        outputWs.cell(row=i, column=11, value=lines[6].strip())
        outputWs.cell(row=i, column=12, value=lines[7].strip())
    
        time.sleep(0.5) #API制限対策

    except:
        print('Error')

outputworkbook.save(f"{outputfile}{outputfilename}")
print("Process End")




